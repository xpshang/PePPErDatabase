import datetime
import io
import json
import math
import uuid

import pytz
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.forms import model_to_dict
from django.http import FileResponse, HttpResponse, JsonResponse, request
from django.shortcuts import render, redirect
from django.views import View
import matplotlib.pyplot as plt
import numpy as np
import base64
from io import BytesIO
import openpyxl
from matplotlib.ticker import MultipleLocator
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from DataApp import tasplot
from decimal import Decimal
from DataApp.models import DataModel, DataColumnName, LiteratureModel, UserUploadModel, HistoryModel
import matplotlib
matplotlib.use("agg")
from django.views import View
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
london_tz = pytz.timezone('Europe/London')
from matplotlib.ticker import FixedLocator
#登录权限保护
class LoginRequiredView(View):
    @classmethod
    def as_view(cls, **initkwargs):
        view = super(LoginRequiredView, cls).as_view()
        return login_required(view)

# Create your views here.
class HomeView(View):
    def get(self,request):
        nav="home"
        return render(request, 'home.html', locals())

#下载页面
class DownloadDataView(View):
    def get(self,request):
        nav="download"
        # 查询下载历史记录
        row=DataModel.objects.last()


        return render(request,'download.html',locals())
# 绘图页面
class DataPlotterView(View):
    def get(self,request):
        # 获取传入参数
        type = request.GET.get("type")
        png=request.GET.get("png")

        st=request.GET.get("st")
        nav="plot"
        # 查询数据
        rows=DataColumnName.objects.filter(cate=1)
        data_rows=DataColumnName.objects.filter(cate=2)
        return render(request,'data_plotter.html',locals())
    # 绘制图标
    def post(self,request):
        nav="plot"
        # 获取要查询的列，已经筛选条件和x和y轴的值
        filed=request.POST.get("m_filed")
        value=request.POST.get("m_value")
        x=request.POST.get("m_x")
        y=request.POST.get("m_y")

        # 调用绘制方法
        src=self.draw(filed,value,x,y,request.user.id)

        return render(request, 'data_plotter_result.html', locals())

    def draw(self,filed,value,x,y,user_id=None):
        # 拆分数据
        valuelist = value.split("__a__")
        plt.figure()
        # 如果x,y等于下方两值，则绘制分区图
        if x.strip() == "bulk_sio2" and y.strip() == "bulk_na2o":
            tasplot.add_LeMaitre_fields(plt)  # add TAS fields to plot
        #     定义散点样式
        markers = ['.','v','^','<','>','s','p','*','h','+','x','D','d','_']
        index=0
        # 循环遍历，将每一种类别散点分别绘制到图上
        for value in valuelist:
            # 取模获取不同的样式
            marker=markers[index%len(markers)]
            index+=1
            # 查询数据
            dataModels = DataModel.objects.filter(**{filed: value}).all()

            xlist = []
            ylist = []
            # 循环遍历，处理不符合条件的数据
            for i in dataModels:
                i = i.__dict__
                if i[x] is not None and i[x].strip() != "" and i[y] is not None and i[y].strip() != "":
                    try:

                        xlist.append(float(i[x]))
                        ylist.append(float(i[y]))
                    except:
                        continue
            xlist = np.array(xlist)
            ylist = np.array(ylist)

            #绘制
            plt.scatter(xlist, ylist,label=value,marker=marker,s=5)
        xInfo = DataColumnName.objects.filter(name=x).first()
        yInfo = DataColumnName.objects.filter(name=y).first()
        plt.xlabel(xInfo.desc)  # 设置x坐标轴的名称
        plt.ylabel(yInfo.desc)  # 设置x坐标轴的名称
        # sio = BytesIO()


        name=str(uuid.uuid4())+".png"
        url="/static/img/"+name

        ncol=1
        #设置条例
        lg=plt.legend(fontsize='7',bbox_to_anchor=(1.05, 1.0), loc='upper left',ncol=ncol)
        plt.tight_layout()
        #保存图片
        plt.savefig("./static/img/"+name, dpi=300,
            format='png',
            bbox_extra_artists=(lg,),
            bbox_inches='tight')
        # 如果登录装填，则保存历史记录

        # plt.savefig(sio, format='png', dpi=500)
        # sio.seek(0)
        # plt.close()
        return url
# 动态获取每一列的分类和数据数量
class DataPlotFieldListView(View):
    def get(self,request):
        filed=request.GET.get("filed")
        distinct_list=DataModel.objects.values(filed).annotate(count=Count("id"))
        items=[]
        for i in distinct_list:
            items.append({"name":i[filed],"count":i['count']})
        return JsonResponse({"code":1,"message":"ok","data":items})


# 下载功能
class ExportDataView(View):
    def get(self,request):
        params=dict(request.GET)
        querys={}
        name=request.GET.get("name",None)
        if name is None or name=="":
            name="download"
        for key,value in params.items():
            if key=="name":
                continue
            querys[key]=value[0]
        # 构建excel
        workbook=Workbook()
        worksheet=workbook.active
        columns=DataColumnName.objects.order_by("id")
        headers=[]
        keys=[]
        # 添加表头
        for i in columns:
            headers.append(i.desc)
            keys.append(i.name)
        worksheet.append(headers)
        header_cell_range = worksheet['A1:Y1']

        # 设置字体为加粗
        font = Font(bold=True)

        # 应用字体样式到表头单元格
        for cell in header_cell_range[0]:
            cell.font = font
        # 特殊处理非experimental的方法：
        if name=="not_experimental":
            datamodels=DataModel.objects.filter(~Q(**querys)).all()
        elif name=="all":
            datamodels=DataModel.objects.all()
        else:
            datamodels=[]
            for key,value in querys.items():
                vl=value.split("__a__")
                if len(vl)==0:
                    datamodels= DataModel.objects.all()
                else:
                    for i in vl:
                        datamodelsdetails = DataModel.objects

                        datamodelsdetails=datamodelsdetails.filter(**{key:i}).all()
                        datamodels.extend(datamodelsdetails)
        #将数据塞入表内
        index=0
        for i in datamodels:
            index+=1
            i=i.__dict__
            line=[]
            for x in keys:
                if x not in i.keys():
                    line.append("")
                elif i[x] is not None:
                    line.append(i[x])
                else:
                    line.append("")

            worksheet.append(line)
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # 调整列宽度的系数
            worksheet.column_dimensions[column_letter].width = adjusted_width
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        #保存exel类型记录
        save_name = str(uuid.uuid4()) + ".xlsx"
        url = "/static/file/" + save_name
        workbook.save("./static/file/"+save_name)
        if request.user.id is not None:
            today=datetime.datetime.now(london_tz)
            HistoryModel.objects.create(url=url, user_id=request.user.id,type=1,name=save_name,crt_date=today)

        # 创建文件响应并设置文件类型和名称
        response = FileResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.xlsx')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(name)

        return response
from openpyxl import load_workbook

# 发送邮件
def sendEmail(account,name,filename):

    # 发件人邮箱和密码（需要开启SMTP服务，并使用授权码登录）
    sender_email = "pepperdatabase@gmail.com"
    sender_password = "tbdylyohhcfjhxmx"

    # 收件人邮箱
    recipient_email = "qu22346@bristol.ac.uk"

    # 构建邮件内容
    msg = MIMEMultipart()
    # msg['From'] = sender_email
    msg['From'] = name
    msg['To'] = recipient_email
    msg['Subject'] = "New Data File Submission - PePPEr Website"

    body="""<html>
<head></head>
<body><p><strong>Dear Administrator,</strong></p>

<div style="margin-top: 20px">We would like to inform you that a new data file has been submitted to the PePPEr website and is pending your review.</div>

<div style="margin-top: 20px">Submission Details:</div>
<div>- User Account: {}</div>
<div>- Username: {}</div>
<div>- File Name: {}</div>
<div>- Submission Time: {}</div>

<div style="margin-top: 20px">
We appreciate your dedication and prompt response to this submission.</div>
<div>Thank you.</div>

<div style="margin-top: 20px">Best regards,</div>
<div>The PePPEr Website</div></body>
</html>""".format(account,name,filename,datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # 邮件正文
    msg.attach(MIMEText(body, 'html'))

    # 连接Gmail的SMTP服务器
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        text = msg.as_string()

        # 发送邮件
        server.sendmail(sender_email, recipient_email, text)
        server.quit()
        print("Email sent successfully!")
    except Exception as e:
        print("Failed to send email. Error: ", str(e))

#上传页面
class UploadView(View):
    def get(self, request):
        nav = "upload"
        columns = DataColumnName.objects.order_by("id")

        return render(request, 'upload.html', locals())
    def post(self, request):
        userid=-1
        if request.user.id is not None:
            userid=request.user.id
        #     创建上传记录
        today = datetime.datetime.now(london_tz)
        uum=UserUploadModel.objects.create(user_id=userid,url=request.FILES["file"],upload_date=today)
        #发送邮件
        sendEmail(request.user.email,request.user.username,request.FILES['file'].name)
        nav = "upload"
        columns = DataColumnName.objects.order_by("id")
        return JsonResponse({"code":200})




#工具页面
class ToolView(View):
    def get(self,request):
        nav='tool'
        return render(request,'tool.html',locals())

    def post(self,request):
        # 加载前端传入的数据
        data=request.body.decode("utf-8")
        json_data=json.loads(data)
        # data=request.data.getlist("data",None)

        json_data=json_data['data']
        type=request.GET.get("type")

        # 按照数据公式进行计算
        if len(json_data) !=0:
            moment_1st = 0
            moment_2nd = 0
            moment_3rd = 0
            b_sum = 0
            c_col=0

            datas={"vf":[],"eq":[],"log":[],"log_2":[],"log_2_d":[],"sum":[]}
            for i in json_data:
                c_col+=Decimal(i["vf"])
                log_2=self.log(Decimal(i['eq']),2)
                log_2_d=-log_2
                log=self.log(Decimal(i['eq']))
                datas['vf'].append(float(Decimal(i["vf"])))
                datas['eq'].append(float(Decimal(i["eq"])))
                datas['log'].append(log)
                datas['log_2'].append(log_2)
                datas['log_2_d'].append(log_2_d)
                datas['sum'].append(float(c_col))
                b_sum+=Decimal(i['vf'])
                value=Decimal(i['eq'])*Decimal(i['vf'])
                moment_1st+=value
                value=Decimal(i['eq'])*value
                moment_2nd+=value
                value=Decimal(i['eq'])*value
                moment_3rd+=value

            moment_1st=moment_1st/b_sum
            moment_2nd=moment_2nd/b_sum

            moment_3rd=moment_3rd/b_sum

            final=round(moment_1st*moment_2nd/moment_3rd,6)
            # 如果是不是下载，分别绘制四个图返回
            if type !="download":

                g1 = self.plot(datas["eq"], datas['sum'], "Equivalent diameter (mm)", "Cumulative fraction", "Equivalent diameter (mm) vs Cumulative fraction",1)
                g2 = self.plot(datas["log_2_d"], datas['sum'], "Equivalent diameter (phi)", "Cumulative fraction", "Equivalent diameter (phi) vs Cumulative fraction",2)
                g3 = self.plot(datas["eq"], datas['vf'], "Log equivalent diameter (mm)", "Volume fraction", "Log equivalent diameter (mm) vs Volume fraction",3)
                g4 = self.plot(datas["eq"], datas['sum'], "Log equivalent diameter (mm)", "Cumulative fraction", "Log equivalent diameter (mm) vs Cumulative fraction",4)

                return JsonResponse({"code": 1, "message": "ok", "data": {"final": '%.6f' % final,
                                                                          "graph": {"g1": g1, "g2": g2, "g3": g3,
                                                                                    "g4": g4}}})
            else:
                #如果是下载，绘制一个大图
                plt.figure(figsize=(20,16))
                self.plotAll(datas["eq"], datas['sum'], "Equivalent diameter (mm)", "Cumulative fraction",
                               "Equivalent diameter (mm) vs Cumulative fraction",1)
                self.plotAll(datas["log_2_d"], datas['sum'], "Equivalent diameter (phi)", "Cumulative fraction",
                               "Equivalent diameter (phi) vs Cumulative fraction",2)
                self.plotAll(datas["eq"], datas['vf'], "Log equivalent diameter (mm)", "Volume fraction",
                               "Log equivalent diameter (mm) vs Volume fraction",3)
                self.plotAll(datas["eq"], datas['sum'], "Log equivalent diameter (mm)", "Cumulative fraction",
                               "Log equivalent diameter (mm) vs Cumulative fraction",4)
                plt.subplots_adjust(left=0.1, right=0.9, bottom=0.1, top=0.9, wspace=0.3, hspace=0.3)
                # plt.show()

                # # 调整子图之间的间距
                # plt.tight_layout()
                buffer = io.BytesIO()
                # name="/static/img/{}.png".format(str(uuid.uuid4()))
                myname=str(uuid.uuid4())
                name="/static/img/{}.png".format(myname)
                plt.savefig("."+name, format='png')
                today = datetime.datetime.now(london_tz)
                if request.user.id is not None:
                    HistoryModel.objects.create(url=name, user_id=request.user.id, type=0, name="{}.png".format(myname),crt_date=today)

                plt.close()
                # response = HttpResponse(buffer, content_type='image/png')
                # response['Content-Disposition'] = 'attachment; filename=plot.png'
                return JsonResponse({"code":200,"data":{"img":name}})

        else:
            final=0
            g1=None
            g2=None
            g3=None
            g4=None


            return JsonResponse({"code":1,"message":"ok","data":{"final":'%.6f'%final,"graph":{"g1":g1,"g2":g2,"g3":g3,"g4":g4}}})

    def log(self,x,base=None):
        if base is not None:
            return math.log(x, base)
        return math.log10(x)
    def plot(self,x,y,x_label,y_label,title,type):
        plt.figure(figsize=(9,6))
        # fig ,ax =plt.subplots()
        xlabel_font = {
            # 'fontsize': rcParams['axes.titlesize'], # 设置成和轴刻度标签一样的大小
            'fontsize': 16,

        }
        # 在这里添加绘图的逻辑
        # ...
        plt.xlabel(x_label,fontdict=xlabel_font,labelpad=12)
        plt.ylabel(y_label,fontdict=xlabel_font,labelpad=12)
        plt.title(title,fontsize=20,pad=15)

        if type==1:
            plt.plot(x, y, marker="o", linestyle="-", color="#5D5968")
        if type==2:
            plt.plot(x, y, marker="o", linestyle="-", color="#5D5968")
            plt.gca().invert_xaxis()
        if type==3:

            plt.semilogx(x,y,marker="o", linestyle="-", color="#5D5968")
            plt.xticks([0.001,0.01,0.1,1,10],['0.001','0.01','0.1','1','10'])
        if type == 4:
            plt.semilogx(x, y, marker="o", linestyle="-", color="#5D5968")
            plt.xticks([0.001,0.01, 0.1, 1, 10], ['0.001','0.01', '0.1', '1', '10'])



        # 存入buffer中
        buffer1 = BytesIO()
        plt.savefig(buffer1, format='png')
        buffer1.seek(0)
        image_png1 = buffer1.getvalue()
        buffer1.close()

        graphic1 = base64.b64encode(image_png1).decode('utf-8')
        return graphic1
    def plotAll(self,x,y,x_label,y_label,title,index):
        print(index)
        xlabel_font = {
            # 'fontsize': rcParams['axes.titlesize'], # 设置成和轴刻度标签一样的大小
            'fontsize': 18,

        }

        plt.subplot(2, 2, index)
        # 在这里添加绘图的逻辑
        # ...
        plt.xlabel(x_label,fontdict=xlabel_font,labelpad=15)
        plt.ylabel(y_label,fontdict=xlabel_font,labelpad=15)
        plt.title(title,fontsize=22,pad=20)
        if index == 1:
            plt.plot(x, y, marker="o", linestyle="-", color="#5D5968")
        if index == 2:
            plt.plot(x, y, marker="o", linestyle="-", color="#5D5968")
            plt.gca().invert_xaxis()
        if index==3:

            plt.semilogx(x,y,marker="o", linestyle="-", color="#5D5968")
            plt.xticks([0.001,0.01,0.1,1,10],['0.001','0.01','0.1','1','10'])
        if index == 4:
            plt.semilogx(x, y, marker="o", linestyle="-", color="#5D5968")
            plt.xticks([0.001,0.01, 0.1, 1, 10], ['0.001','0.01', '0.1', '1', '10'])





# finder页面
class FinderView(View):
    def get(self,request):
        nav='finder'
        queryset = LiteratureModel.objects.all()
        type=request.GET.get("type")
        if type=="download":
            # 如果是下载功能，则查出来塞入excel然后返回
            workbook = Workbook()
            worksheet = workbook.active
            line = ["Publication", "Pub.Year","DOI", "Title"]
            worksheet.append(line)
            for i in queryset:
                line = [i.publication,i.pub_year,i.doi,i.title]
                worksheet.append(line)

            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # 创建文件响应并设置文件类型和名称
            response = FileResponse(output,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.xlsx')
            response['Content-Disposition'] = 'attachment; filename="literature.xlsx"'

            return response
        return render(request,'finder.html',locals())
from django.contrib.auth.models import User

class AdminReviewView(LoginRequiredView):
    def get(self,request):
        uum=UserUploadModel.objects.filter(status=0).all()
        rows=[]
        for i in uum:
            row={}
            row['data']=i
            u=User.objects.filter(id=i.user_id).first()
            row['user']=u
            rows.append(row)
        return render(request,"review.html",locals())

# 管理员审核
class ReviewSubmitView(LoginRequiredView):
    def get(self, request):
        id = request.GET.get("id")
        type = request.GET.get("type")
        # 拒绝业务
        if type=="reject":

            # 直接吸怪状态为拒绝
            uu=UserUploadModel.objects.filter(id=id).first()
            uu.status=-1
            uu.save()


        return redirect("/admin/DataApp/useruploadmodel/")
#历史记录查询
class HistoryView(LoginRequiredView):
    def get(self,request):
        nav='history'

        rows=HistoryModel.objects.filter(user_id=request.user.id).order_by("-crt_date")
        return render(request,'history.html',locals())


from django.http import HttpResponse
from django.conf import settings
import os
from io import BytesIO
from PIL import Image
# 获取图片
def get_image(request):

    # 获取图片路径
    png=request.GET.get("png")
    # 打开图片并转换为IO流
    with open("."+png, "rb") as f:
        image_content = f.read()
    if request.user.id is not None:
        today = datetime.datetime.now(london_tz)
        HistoryModel.objects.create(url=png, user_id=request.user.id, type=0, name=png.split("/")[-1],crt_date=today)
    # 创建IO流对象
    io_stream = BytesIO(image_content)

    # 打开图片
    image = Image.open(io_stream)

    # 将图片转换为IO流
    image_io = BytesIO()
    image.save(image_io, format='PNG')  # 这里假设转换为PNG格式，您可以根据需求选择其他格式

    # 将IO流的指针移动到开始位置
    image_io.seek(0)

    # 返回IO流给前端
    response = HttpResponse(image_io, content_type='image/png')
    response['Content-Disposition'] = 'attachment; filename="plot.png"'

    return response
class AdminUploadView(View):
    def get(self, request):
        pass
# 添加数据
def addDataNow(request):

    id=request.GET.get("id")
    uum= UserUploadModel.objects.filter(id=id).first()
    # 读取文件，
    excel_file=request.FILES["file"]

    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    dcns = DataColumnName.objects.all()
    dicts = {}
    for i in dcns:
        dicts[i.desc] = i.name
    headers_value = [cell.value for cell in worksheet[1]]
    headers = []

    for i in headers_value:
        if i is None:
            continue
        headers.append(dicts[i])
    data = []

    for row in worksheet.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        row_data = dict(zip(headers, values))
        data.append(row_data)
        #     批量存入数据库
    model_objects = [DataModel(**i) for i in data]
    DataModel.objects.bulk_create(model_objects)
    uum.status=1
    uum.save()

    return JsonResponse({"code":200})
class UploadListView(View):
    def get(self,request):
        rows2=UserUploadModel.objects.filter(user_id=request.user.id).all().order_by("-upload_date")
        rows=[]
        #   user_id=models.IntegerField(default=0)
        #     status=models.IntegerField(default=0) #状态：0：待审核，1审核通过,-1：驳回
        #     url=models.FileField(default=2048,upload_to='static/file/')
        #     upload_date=models.DateTimeField(default=datetime.datetime.now(),verbose_name="verbose
        for i in rows2:
            print(i.url.name)
            rows.append({"status":i.status,"url":i.url.name.split("/")[-1],"upload_date":i.upload_date})
        return render(request,'upload_history.html',locals())
